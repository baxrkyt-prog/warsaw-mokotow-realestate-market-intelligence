"""
collectors.transactions.nbp — NBP BaRN: ceny mieszkań (ofertowe + transakcyjne).

Źródło: https://static.nbp.pl/dane/rynek-nieruchomosci/ceny_mieszkan.xlsx
  - kwartalne, od III kw. 2006
  - poziom MIASTA (Warszawa = całe miasto, nie dzielnica!)
  - arkusze: 'Rynek pierwotny', 'Rynek wtórny'
  - bloki kolumn: Ceny ofertowe | Ceny transakcyjne | indeksy hedoniczne

Zapis do transaction_market_snapshots z:
  district='Warszawa', district_norm=NULL (celowo — to poziom miasta,
  nie wolno mieszać z agregatami dzielnicowymi), period_type='quarterly',
  source='nbp_barn'.

Używać w analizach wyłącznie jako benchmark referencyjny miasta.
"""

from __future__ import annotations

import argparse
import re
import urllib.request
from datetime import datetime, timezone

from collectors.base import Collector, CollectorResult
from collectors.registry import register


NBP_URL = "https://static.nbp.pl/dane/rynek-nieruchomosci/ceny_mieszkan.xlsx"
TIMEOUT_S = 60

SHEETS = {
    "Rynek pierwotny": "primary",
    "Rynek wtórny":    "secondary",
}

QUARTER_RE = re.compile(r"^(I{1,3}|IV)\s+(\d{4})$")
ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4}


def quarter_to_date(label: str) -> str | None:
    """'III 2024' → '2024-09-30' (koniec kwartału)."""
    m = QUARTER_RE.match(str(label).strip())
    if not m:
        return None
    q = ROMAN[m.group(1)]
    year = int(m.group(2))
    month_end = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}[q]
    return f"{year}-{month_end}"


def _find_city_columns(rows: list, city: str) -> dict[str, int]:
    """Lokalizuje kolumny miasta w blokach 'Ceny ofertowe' / 'Ceny transakcyjne'.

    rows[3] = nagłówki bloków (komórki scalone → tylko pierwsza kolumna bloku ma tekst)
    rows[6] = nazwy miast
    Zwraca {'offer': col_idx | None, 'transaction': col_idx | None}
    """
    blocks_row = rows[3]
    cities_row = rows[6]

    # Zbuduj zakresy bloków: idx kolumny → nazwa bloku
    block_starts = [(i, str(v)) for i, v in enumerate(blocks_row) if v]
    out = {"offer": None, "transaction": None}

    for bi, (start, bname) in enumerate(block_starts):
        end = block_starts[bi + 1][0] if bi + 1 < len(block_starts) else len(cities_row)
        bl = bname.lower()
        key = None
        if "ofertowe" in bl:
            key = "offer"
        elif "transakcyjne" in bl:
            key = "transaction"
        if key is None:
            continue
        for col in range(start, end):
            cv = cities_row[col]
            if cv and city.lower() in str(cv).lower():
                out[key] = col
                break
    return out


@register
class NbpBarnCollector(Collector):
    source = "nbp_barn"
    kind = "aggregates"
    schema_version = 1
    description = "NBP BaRN — kwartalne ceny mieszkań Warszawa (transakcyjne + ofertowe)"

    @classmethod
    def add_cli_args(cls, parser: argparse.ArgumentParser) -> None:
        parser.add_argument("--file", default=None,
                            help="Lokalny XLSX (pomija pobieranie z NBP)")
        parser.add_argument("--city", default="Warszawa")
        parser.add_argument("--dry-run", action="store_true")

    def run(self, **kwargs) -> CollectorResult:
        from openpyxl import load_workbook
        from database import get_conn, upsert_transaction_market_snapshot

        result = CollectorResult(source=self.source, kind="aggregates")
        city = kwargs.get("city") or "Warszawa"
        dry = kwargs.get("dry_run", False)

        # 1. Pobierz / otwórz plik
        path = kwargs.get("file")
        if not path:
            result.extras["url_scraped"] = NBP_URL
            try:
                req = urllib.request.Request(
                    NBP_URL, headers={"User-Agent": "Mozilla/5.0"})
                data = urllib.request.urlopen(req, timeout=TIMEOUT_S).read()
            except Exception as e:
                result.status = "error"
                result.error_msg = f"download failed: {e}"
                return result
            import tempfile
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.write(data)
            tmp.close()
            path = tmp.name
        else:
            result.extras["url_scraped"] = path

        # 2. Parsuj arkusze
        wb = load_workbook(path, read_only=True, data_only=True)
        now = datetime.now(timezone.utc).isoformat()
        new_count = 0

        with get_conn() as conn:
            for sheet_name, market_type in SHEETS.items():
                if sheet_name not in wb.sheetnames:
                    result.reject(f"missing_sheet:{sheet_name}")
                    continue
                rows = list(wb[sheet_name].iter_rows(values_only=True))
                cols = _find_city_columns(rows, city)
                if cols["transaction"] is None and cols["offer"] is None:
                    result.reject(f"city_not_found:{sheet_name}")
                    continue

                for r in rows[7:]:
                    label = r[0]
                    snap_date = quarter_to_date(label) if label else None
                    if not snap_date:
                        continue
                    tx_val = r[cols["transaction"]] if cols["transaction"] is not None else None
                    if tx_val is None or not isinstance(tx_val, (int, float)):
                        continue
                    result.records_in += 1
                    if dry:
                        continue
                    upsert_transaction_market_snapshot(conn, {
                        "snapshot_date": snap_date,
                        "period_type": "quarterly",
                        "source": self.source,
                        "property_type": "residential",
                        "market_type": market_type,
                        "district": city,
                        "subdistrict": None,
                        "district_norm": None,  # poziom miasta — celowo NULL
                        "transaction_count": None,  # NBP nie publikuje liczby w tym pliku
                        "median_price_per_m2": None,
                        "average_price_per_m2": round(float(tx_val), 2),
                        "transaction_volume_pln": None,
                        "imported_at": now,
                    })
                    new_count += 1

        result.records_new = new_count
        result.extras["city"] = city
        result.extras["property_type"] = "residential"
        return result
