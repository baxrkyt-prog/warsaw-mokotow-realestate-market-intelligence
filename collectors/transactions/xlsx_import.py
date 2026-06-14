"""
collectors.transactions.xlsx_import — generyczny import XLSX → transactions.
Używa openpyxl (zazwyczaj już zainstalowany jako transient dep pandas).
"""

from __future__ import annotations

import argparse
from pathlib import Path

from collectors.base import Collector, CollectorResult
from collectors.registry import register
from collectors.transactions._tabular import ImportConfig, import_rows


def _iter_xlsx_rows(path: str, sheet: str | None, skip_rows: int):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl nie jest dostępne — uruchom: pip install openpyxl"
        ) from e

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    for _ in range(skip_rows):
        next(rows, None)
    header = next(rows, None)
    if not header:
        return
    header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header)]
    for row in rows:
        if row is None or all(c is None or c == "" for c in row):
            continue
        yield dict(zip(header, row))


@register
class XlsxTransactionCollector(Collector):
    source = "xlsx_import"
    kind = "transactions"
    schema_version = 1
    description = "Generic XLSX → transactions (JSON mapping)"

    @classmethod
    def add_cli_args(cls, parser: argparse.ArgumentParser) -> None:
        parser.add_argument("--file", required=True, help="Plik XLSX")
        parser.add_argument("--mapping", required=True, help="JSON z konfigiem importu")
        parser.add_argument("--sheet", default=None,
                            help="Nazwa arkusza (domyślnie z mappingu, fallback: pierwszy)")
        parser.add_argument("--dry-run", action="store_true")

    def run(self, **kwargs) -> CollectorResult:
        cfg = ImportConfig.from_file(kwargs["mapping"])
        result = CollectorResult(source=cfg.source, kind="transactions")
        result.extras["url_scraped"] = str(Path(kwargs["file"]).resolve())

        try:
            sheet = kwargs.get("sheet") or cfg.sheet
            rows = _iter_xlsx_rows(kwargs["file"], sheet, cfg.skip_rows)
            import_rows(rows, cfg, result, dry_run=kwargs.get("dry_run", False))
        except FileNotFoundError as e:
            result.status = "error"
            result.error_msg = f"file not found: {e}"
        except Exception as e:
            result.status = "error"
            result.error_msg = str(e)

        return result
