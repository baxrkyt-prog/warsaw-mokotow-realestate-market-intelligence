"""
collectors.transactions.rcn — import wypisu z Rejestru Cen Nieruchomości (RCN).

RCN (dawniej RCiWN) prowadzi starosta / prezydent miasta. Dane NIE są publiczne —
wypis uzyskuje się wnioskiem (formularz P + P5) u BGiK m.st. Warszawy lub przez
geoportal powiatowy. Zobacz: docs/RCN_ACCESS.md.

Ten collector przyjmuje typowy format eksportu RCN (CSV/XLSX) i:
  - auto-wykrywa polskie nagłówki kolumn (data transakcji, cena, powierzchnia, ...)
  - alternatywnie przyjmuje jawny --mapping JSON (jak csv_import)
  - filtruje typ nieruchomości (lokal mieszkalny / użytkowy / grunt)
  - geokoduje adresy przez GUGiK (cache-first) — RCN podaje adresy/obręby
  - zapisuje do `transactions` z source='rcn_<powiat>'

Użycie:
  python -m collectors run rcn --file wypis_rcn.csv --powiat warszawa
  python -m collectors run rcn --file wypis.xlsx --mapping custom.json
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from collectors.base import Collector, CollectorResult
from collectors.registry import register
from collectors.transactions._tabular import ImportConfig, import_rows


# Auto-detekcja nagłówków: docelowe pole → lista wzorców (lowercase substring match)
HEADER_PATTERNS: dict[str, list[str]] = {
    "transaction_date":  ["data transakcji", "data zawarcia", "data aktu", "data umowy"],
    "transaction_price": ["cena transakcyjna", "cena brutto", "cena nieruchomości", "cena [zł]", "cena"],
    "area_m2":           ["powierzchnia lokalu", "pow. lokalu", "powierzchnia użytkowa",
                          "pow. użytkowa", "powierzchnia [m2]", "powierzchnia"],
    "address":           ["adres nieruchomości", "położenie", "adres", "ulica"],
    "subdistrict":       ["obręb ewidencyjny", "obręb", "dzielnica"],
    "rooms":             ["liczba izb", "liczba pokoi", "izby"],
    "floor":             ["kondygnacja", "piętro"],
    "year_built":        ["rok budowy", "rok zakończenia budowy"],
    "source_record_id":  ["numer transakcji", "id transakcji", "lp.", "lp"],
}

# Wartości w kolumnie "rodzaj nieruchomości" wskazujące mieszkania
RESIDENTIAL_MARKERS = ["lokal mieszkalny", "mieszkalny", "mieszkanie"]
PROPERTY_TYPE_HEADERS = ["rodzaj nieruchomości", "typ nieruchomości", "rodzaj", "przedmiot transakcji"]


def detect_mapping(header: list[str]) -> tuple[dict[str, str], str | None]:
    """Z nagłówka pliku buduje mapping {pole_docelowe: kolumna_źródłowa}.
    Zwraca też nazwę kolumny z rodzajem nieruchomości (do filtrowania)."""
    mapping: dict[str, str] = {}
    lower = {h: str(h).strip().lower() for h in header if h}

    for target, patterns in HEADER_PATTERNS.items():
        best = None
        for pat in patterns:  # patterns w kolejności od najbardziej specyficznych
            for col, low in lower.items():
                if pat in low:
                    best = col
                    break
            if best:
                break
        if best:
            mapping[target] = best

    ptype_col = None
    for col, low in lower.items():
        if any(p in low for p in PROPERTY_TYPE_HEADERS):
            ptype_col = col
            break
    return mapping, ptype_col


def _iter_file(path: str, delimiter: str, encoding: str, sheet: str | None):
    """Yield dictów z CSV lub XLSX."""
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return
        header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header)]
        for row in rows:
            if row is None or all(c is None or c == "" for c in row):
                continue
            yield dict(zip(header, row))
    else:
        with open(path, "r", encoding=encoding, newline="") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            yield from reader


@register
class RcnCollector(Collector):
    source = "rcn"
    kind = "transactions"
    schema_version = 1
    description = "Wypis z Rejestru Cen Nieruchomości (auto-detect nagłówków + geocoding)"

    @classmethod
    def add_cli_args(cls, parser: argparse.ArgumentParser) -> None:
        parser.add_argument("--file", required=True, help="Wypis RCN (CSV lub XLSX)")
        parser.add_argument("--powiat", default="warszawa",
                            help="Identyfikator powiatu (część source ID)")
        parser.add_argument("--mapping", default=None,
                            help="Jawny JSON mapping (pomija auto-detekcję)")
        parser.add_argument("--market-type", default=None,
                            choices=["primary", "secondary"],
                            help="Wymuś rynek (RCN zwykle nie rozróżnia)")
        parser.add_argument("--sheet", default=None, help="Arkusz XLSX")
        parser.add_argument("--delimiter", default=";",
                            help="Separator CSV (wypisy RCN zwykle ';')")
        parser.add_argument("--encoding", default="utf-8",
                            help="Kodowanie (starostwa często cp1250)")
        parser.add_argument("--no-geocode", action="store_true")
        parser.add_argument("--geocode-limit", type=int, default=500)
        parser.add_argument("--dry-run", action="store_true")

    def run(self, **kwargs) -> CollectorResult:
        src_id = f"rcn_{kwargs.get('powiat', 'warszawa')}"
        result = CollectorResult(source=src_id, kind="transactions")
        result.extras["url_scraped"] = str(Path(kwargs["file"]).resolve())

        delimiter = kwargs.get("delimiter") or ";"
        encoding = kwargs.get("encoding") or "utf-8"
        sheet = kwargs.get("sheet")

        # 1. Konfig: jawny mapping albo auto-detekcja
        ptype_col = None
        if kwargs.get("mapping"):
            cfg = ImportConfig.from_file(kwargs["mapping"])
            cfg.source = src_id  # wymuszamy spójny source
        else:
            try:
                first = next(_iter_file(kwargs["file"], delimiter, encoding, sheet), None)
            except FileNotFoundError as e:
                result.status = "error"
                result.error_msg = f"file not found: {e}"
                return result
            except UnicodeDecodeError:
                result.status = "error"
                result.error_msg = (f"encoding {encoding} failed — wypisy ze starostw "
                                    "bywają w cp1250, spróbuj --encoding cp1250")
                return result
            if first is None:
                result.status = "error"
                result.error_msg = "pusty plik"
                return result

            columns, ptype_col = detect_mapping(list(first.keys()))
            missing = {"transaction_date", "transaction_price", "area_m2"} - set(columns)
            if missing:
                result.status = "error"
                result.error_msg = (f"auto-detekcja nie znalazła kolumn: {missing}. "
                                    f"Wykryte: {columns}. Podaj jawny --mapping.")
                return result

            result.extras["detected_mapping"] = columns
            cfg = ImportConfig(
                source=src_id,
                property_type="residential",
                market_type=kwargs.get("market_type"),
                default_district="Warszawa",
                default_district_norm=None,
                date_format="%Y-%m-%d",
                delimiter=delimiter,
                encoding=encoding,
                columns=columns,
                id_strategy="source_record_id" if "source_record_id" in columns else "hash",
                geocode=not kwargs.get("no_geocode", False),
                geocode_limit=int(kwargs.get("geocode_limit", 500)),
            )

        # 2. Iteruj z filtrem rodzaju nieruchomości (tylko lokale mieszkalne)
        def filtered_rows():
            for raw in _iter_file(kwargs["file"], delimiter, encoding, sheet):
                if ptype_col:
                    val = str(raw.get(ptype_col) or "").lower()
                    if val and not any(m in val for m in RESIDENTIAL_MARKERS):
                        result.reject("non_residential")
                        continue
                yield raw

        try:
            import_rows(filtered_rows(), cfg, result, dry_run=kwargs.get("dry_run", False))
        except UnicodeDecodeError:
            result.status = "error"
            result.error_msg = (f"encoding {encoding} failed mid-file — spróbuj --encoding cp1250")
        except Exception as e:
            result.status = "error"
            result.error_msg = str(e)
        return result
